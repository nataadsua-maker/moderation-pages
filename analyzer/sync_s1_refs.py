"""Sync the System1 reference tables from Google Sheets → Cloudflare KV.

Two tables, two KV keys:

  s1_buyers    — «Баеры S1»: ник → s1pagid по каждому сорсу, воркспейс, домен.
                 s1pagid НЕЛЬЗЯ вывести из ника: у части баеров написание своё,
                 и выравнивать его нельзя — собьётся подсчёт бонусов.
  s1_segments  — «Сегменты»: домен ленда → сегмент по каждому сорсу + статус домена.

Читаем через экспорт Drive в xlsx — тем же сервисным аккаунтом, что уже тянет
ростер, поэтому Sheets API включать не нужно. Запускается из GitHub Actions.
"""
from __future__ import annotations
import io
import json
import os
import sys
import urllib.request

import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build

SOURCES = ["NB", "FB", "MG", "TB", "SN", "OB", "TT"]

# Человеческие названия из таблицы → id в ClickFlare. Незнакомое имя не роняет
# синк: пишем raw-значение и предупреждение в лог, чтобы было видно при разборе.
WORKSPACE_IDS = {
    "RSOC": "675aba9a03358d00126edba0",
    "RSOC2": "67a21b41c01f0e00125a998d",
    "RSOC BA 2": "67d7f9cb7d0af600127847c4",
    "NataN": "67e14fd10be2a40012f7a5b1",
    "Taboola": "664328a5085f110012c3780c",
    "Anya team": "67922040e505a00012f04a7d",
    "Facebook main Team": "65c337182b5a950012bc5af8",
    "Guide": "6981ffd705eda80012d9766e",
    "Yuri tests": "694168376a7c2f00120921c5",
    "Test_Loans": "699dd1ae003fa200123e388c",
}
DOMAIN_IDS = {
    "trk.irarh.space": "66e1674c061c500013ac4fa8",
    "track.igtsmts.space": "65267476ad587c0012e01af5",
    "trk.gengavox.com": "6981f8145b73280012367f0a",
    "track.find-answer.com": "6a6b2fa45b95250012b5a367",
    "track.evtmtyl.space": "652674f0de827600129fd262",
    "flarequick.com": "652665e70c4b7e0012ebe8b0",  # легаси: новое на нём не запускаем
}
LEGACY_DOMAINS = {"flarequick.com"}


def drive_client():
    sa = json.loads(os.environ["GOOGLE_DRIVE_SA_JSON"])
    print("service account:", sa.get("client_email"))
    creds = service_account.Credentials.from_service_account_info(
        sa, scopes=["https://www.googleapis.com/auth/drive.readonly"],
    )
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def load_book(drive, sheet_id: str, label: str):
    try:
        data = drive.files().export(
            fileId=sheet_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ).execute()
    except Exception as e:  # доступ не выдан / файл удалён — говорим прямо
        sys.exit(f"{label}: не удалось прочитать таблицу {sheet_id}: {e}")
    return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)


def cell(row: tuple, i: int) -> str:
    return str(row[i]).strip() if 0 <= i < len(row) and row[i] is not None else ""


def find_header(rows: list, must_have: list) -> int:
    """Первая строка, где встречаются все нужные заголовки. -1 если нет."""
    for n, row in enumerate(rows):
        low = [str(c or "").strip().lower() for c in row]
        if all(any(m.lower() == c or m.lower() in c for c in low) for m in must_have):
            return n
    return -1


def col_index(header: list, name: str) -> int:
    want = name.strip().lower()
    for i, h in enumerate(header):
        if str(h or "").strip().lower() == want:
            return i
    for i, h in enumerate(header):
        if want in str(h or "").strip().lower():
            return i
    return -1


def parse_buyers(wb) -> dict:
    for name in wb.sheetnames:
        rows = list(wb[name].iter_rows(values_only=True))
        h = find_header(rows, ["Ник (нейминг)"])
        if h < 0:
            continue
        header = [str(c or "") for c in rows[h]]
        ci_nick = col_index(header, "Ник (нейминг)")
        ci_ws = col_index(header, "Воркспейс")
        ci_dom = col_index(header, "Домен по умолчанию")
        out: dict = {}
        for row in rows[h + 1:]:
            nick = cell(row, ci_nick)
            if not nick:
                continue
            pagid = {}
            for s in SOURCES:
                ci = col_index(header, f"s1pagid {s}")
                v = cell(row, ci)
                if v:
                    pagid[s] = v
            ws_name = cell(row, ci_ws)
            dom_name = cell(row, ci_dom)
            if ws_name and ws_name not in WORKSPACE_IDS:
                print(f"  ⚠ {nick}: неизвестный воркспейс «{ws_name}» — id не проставлен")
            if dom_name and dom_name not in DOMAIN_IDS:
                print(f"  ⚠ {nick}: неизвестный домен «{dom_name}» — id не проставлен")
            if dom_name in LEGACY_DOMAINS:
                print(f"  ⚠ {nick}: домен «{dom_name}» легаси, новое на нём не запускаем")
            out[nick.lower()] = {
                "nick": nick,
                "pagid": pagid,
                "workspace": ws_name or None,
                "workspaceId": WORKSPACE_IDS.get(ws_name),
                "domain": dom_name or None,
                "domainId": DOMAIN_IDS.get(dom_name),
            }
        print(f"баеры: вкладка «{name}», строк {len(out)}")
        return out
    sys.exit("баеры: не нашла вкладку с колонкой «Ник (нейминг)»")


def parse_segments(wb) -> dict:
    # В таблице две шапки подряд: первая с колонкой «сегменты» — она нам и нужна,
    # вторая содержит готовые Offer URL и здесь не используется.
    for name in wb.sheetnames:
        rows = list(wb[name].iter_rows(values_only=True))
        h = find_header(rows, ["System1", "сегменты"])
        if h < 0:
            continue
        header = [str(c or "") for c in rows[h]]
        ci_dom = col_index(header, "System1")
        ci_st = col_index(header, "Статус домена")
        out: dict = {}
        for row in rows[h + 1:]:
            dom = cell(row, ci_dom)
            if not dom:
                continue
            if dom.lower().startswith("system1"):  # началась вторая шапка
                break
            segs = {}
            for s in SOURCES:
                ci = col_index(header, s)
                v = cell(row, ci)
                if v and v != "-":
                    segs[s] = v
            if not segs:
                continue
            out[dom.lower()] = {"domain": dom, "status": cell(row, ci_st), "segments": segs}
        active = sum(1 for v in out.values() if "актив" in v["status"].lower())
        print(f"сегменты: вкладка «{name}», доменов {len(out)} (активных {active})")
        return out
    sys.exit("сегменты: не нашла вкладку с колонками «System1» и «сегменты»")


def kv_put(key: str, value: dict) -> None:
    acc = os.environ["CF_ACCOUNT_ID"]
    tok = os.environ["CF_API_TOKEN"]
    ns = os.environ["CF_KV_NAMESPACE"]
    url = f"https://api.cloudflare.com/client/v4/accounts/{acc}/storage/kv/namespaces/{ns}/values/{key}"
    req = urllib.request.Request(
        url, data=json.dumps(value, ensure_ascii=False).encode(), method="PUT",
        headers={"Authorization": f"Bearer {tok}", "Content-Type": "text/plain"},
    )
    resp = json.loads(urllib.request.urlopen(req).read().decode())
    print(f"KV {key}: success={resp.get('success')}")
    if not resp.get("success"):
        sys.exit(f"KV write failed for {key}: {resp}")


def main() -> None:
    drive = drive_client()
    buyers = parse_buyers(load_book(drive, os.environ["S1_BUYERS_SHEET_ID"], "баеры"))
    segments = parse_segments(load_book(drive, os.environ["S1_SEGMENTS_SHEET_ID"], "сегменты"))
    kv_put("s1_buyers", buyers)
    kv_put("s1_segments", segments)


if __name__ == "__main__":
    main()
