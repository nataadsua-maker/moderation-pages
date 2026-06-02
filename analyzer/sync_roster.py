"""Sync the moderation-service buyer roster from the team Google Sheet → Cloudflare KV.

Reads the «Сотрудники» tab of «Команда_Nataliia», builds a map
telegram_handle_lower → {trackerNick, fio} (excluding dismissed employees), and
PUTs it to KV key `roster`. Runs on a schedule from GitHub Actions.

Uses the Drive API export (xlsx) — the same service account already has Drive
access for video archiving, so the Sheets API does not need to be enabled.
"""
from __future__ import annotations
import io
import json
import os
import sys
import urllib.request

import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build


def find_col(header: list, substr: str) -> int:
    for i, h in enumerate(header):
        if h and substr.lower() in str(h).lower():
            return i
    return -1


def main() -> None:
    sa = json.loads(os.environ["GOOGLE_DRIVE_SA_JSON"])
    print("service account:", sa.get("client_email"))
    creds = service_account.Credentials.from_service_account_info(
        sa, scopes=["https://www.googleapis.com/auth/drive.readonly"],
    )
    drive = build("drive", "v3", credentials=creds, cache_discovery=False)

    sid = os.environ["TEAM_SHEET_ID"]
    data = drive.files().export(
        fileId=sid,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ).execute()

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)

    # Pick the «Сотрудники» sheet by name; fallback: the sheet whose header has
    # ФИО + Ник в трекере + Телега.
    ws = None
    for name in wb.sheetnames:
        if "сотрудник" in name.lower():
            ws = wb[name]
            break
    chosen = None
    if ws is not None:
        chosen = ws.title
    rosters_candidates = []
    for name in wb.sheetnames:
        sheet = wb[name]
        first = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        hdr = [str(c or "") for c in first]
        if find_col(hdr, "Ник в трекере") >= 0 and find_col(hdr, "Телега") >= 0 and find_col(hdr, "ФИО") >= 0:
            rosters_candidates.append(name)
    if ws is None:
        if not rosters_candidates:
            sys.exit("could not find Сотрудники sheet (no ФИО/Ник/Телега columns)")
        chosen = rosters_candidates[0]
        ws = wb[chosen]
    print("tab:", chosen)

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit("empty sheet")
    hdr = [str(c or "") for c in rows[0]]
    ci_fio = find_col(hdr, "ФИО")
    ci_nick = find_col(hdr, "Ник в трекере")
    ci_tg = find_col(hdr, "Телега")
    ci_status = find_col(hdr, "Стату")
    if min(ci_fio, ci_nick, ci_tg) < 0:
        sys.exit(f"required columns not found (ФИО={ci_fio} Ник={ci_nick} Телега={ci_tg})")

    def cell(row: tuple, i: int) -> str:
        return str(row[i]).strip() if 0 <= i < len(row) and row[i] is not None else ""

    roster: dict[str, dict] = {}
    fired = 0
    for row in rows[1:]:
        tg = cell(row, ci_tg)
        if not (tg.startswith("@") and len(tg) > 1):
            continue
        if "увол" in cell(row, ci_status).lower():
            fired += 1
            continue
        key = tg.lower().lstrip("@").strip()
        if key:
            roster[key] = {"trackerNick": cell(row, ci_nick) or None, "fio": cell(row, ci_fio) or None}
    print(f"roster entries: {len(roster)} (skipped fired: {fired})")

    acc = os.environ["CF_ACCOUNT_ID"]
    tok = os.environ["CF_API_TOKEN"]
    ns = os.environ["CF_KV_NAMESPACE"]
    url = f"https://api.cloudflare.com/client/v4/accounts/{acc}/storage/kv/namespaces/{ns}/values/roster"
    req = urllib.request.Request(
        url, data=json.dumps(roster, ensure_ascii=False).encode(), method="PUT",
        headers={"Authorization": f"Bearer {tok}", "Content-Type": "text/plain"},
    )
    resp = json.loads(urllib.request.urlopen(req).read().decode())
    print("KV write success:", resp.get("success"))
    if not resp.get("success"):
        sys.exit(f"KV write failed: {resp}")


if __name__ == "__main__":
    main()
